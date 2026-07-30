"""Spreadsheet ingestion — XLSX / XLS / ODS / CSV → structured Academy knowledge.

Never stores raw workbooks. Extracts formulas, variables, templates, ratio names.
"""

from __future__ import annotations

import hashlib
import io
import re
from typing import Any

from academy.books.classify import classify_academy
from academy.books.copyright import scrub_definition, scrub_explanation
from academy.books.schema import BookConcept, BookMeta, FormulaObject, FrameworkObject


_RATIO_HINTS = {
    "roe": ("ROE", "Net Income / Equity", "Return on equity"),
    "roa": ("ROA", "Net Income / Assets", "Return on assets"),
    "roce": ("ROCE", "EBIT / Capital Employed", "Return on capital employed"),
    "roic": ("ROIC", "NOPAT / Invested Capital", "Return on invested capital"),
    "wacc": ("WACC", "E/V*re + D/V*rd*(1-t)", "Weighted average cost of capital"),
    "fcf": ("Free Cash Flow", "CFO - Capex", "Free cash flow"),
    "dcf": ("DCF", "Σ CF_t/(1+r)^t + TV/(1+r)^n", "Discounted cash flow value"),
    "capm": ("CAPM", "rf + beta*ERP", "Capital asset pricing model"),
    "ev/ebitda": ("EV/EBITDA", "Enterprise Value / EBITDA", "Enterprise multiple"),
    "pe": ("P/E", "Price / EPS", "Price to earnings"),
    "pb": ("P/B", "Price / Book", "Price to book"),
    "dividend": ("Dividend Yield", "DPS / Price", "Income yield"),
    "terminal": ("Terminal Value", "FCF_(n+1)/(r-g)", "Continuing value"),
    "nopat": ("NOPAT", "EBIT*(1-t)", "Net operating profit after tax"),
    "eva": ("EVA", "NOPAT - WACC*Capital", "Economic value added"),
}

_FORMULA_CELL_RE = re.compile(r"^=.+")


def ingest_spreadsheet(
    *,
    filename: str,
    content_bytes: bytes,
    title: str | None = None,
    store: Any = None,
) -> dict[str, Any]:
    from academy.books.graph import rebuild_graph
    from academy.books.ingest import ensure_seeded
    from academy.books.store import get_books_store

    store = store or get_books_store()
    ensure_seeded(store)

    parsed = parse_spreadsheet(filename=filename, content_bytes=content_bytes)
    if not parsed.get("ok"):
        return {"ok": False, "reason": parsed.get("reason") or "parse_failed", "kind": "spreadsheet"}

    book_id = _sheet_id(title or filename)
    meta = BookMeta(
        book_id=book_id,
        title=title or Path_stem(filename),
        authors=["Spreadsheet Model"],
        subject="Financial Model / Dataset",
        topics=parsed.get("topics") or ["spreadsheet", "model"],
        difficulty="advanced",
        source_format=parsed.get("format") or "xlsx",
        academies=parsed.get("academies") or ["valuation"],
    )
    store.upsert_book(meta)

    concepts: list[BookConcept] = []
    formulas: list[FormulaObject] = []
    frameworks: list[FrameworkObject] = []

    for var in parsed.get("variables") or []:
        name = str(var.get("name") or "").strip()
        if len(name) < 2:
            continue
        academy = classify_academy(name)
        c = BookConcept(
            concept_id=_cid("ssc", name, book_id),
            title=name[:80],
            definition=scrub_definition(var.get("meaning") or f"Spreadsheet variable / input: {name}"),
            explanation=scrub_explanation(var.get("meaning") or f"Model input used in {meta.title}"),
            academy=academy,
            source_book_id=book_id,
            source_chapter=var.get("sheet"),
            confidence=0.7,
            kind="terminology",
            linked_formulas=[],
        )
        store.upsert_concept(c)
        concepts.append(c)

    for f in parsed.get("formulas") or []:
        fo = FormulaObject(
            formula_id=_cid("ssf", f.get("name") or "formula", book_id),
            name=str(f.get("name") or "Model Formula")[:80],
            expression=scrub_definition(str(f.get("expression") or ""))[:160],
            explanation=scrub_explanation(f.get("meaning") or "Spreadsheet calculation logic"),
            variables=dict(f.get("variables") or {}),
            use_cases=list(f.get("use_cases") or ["valuation model", "ratio template"]),
            academy=classify_academy(str(f.get("name") or "")),
            source_book_id=book_id,
            source_chapter=f.get("sheet"),
            confidence=float(f.get("confidence") or 0.75),
        )
        store.upsert_formula(fo)
        formulas.append(fo)
        meta.academies = sorted(set(meta.academies + [fo.academy]))

    # Template → soft framework
    for tmpl in parsed.get("templates") or []:
        fw = FrameworkObject(
            framework_id=_cid("ssfw", tmpl.get("name") or "template", book_id),
            name=str(tmpl.get("name") or "Model Template")[:80],
            purpose=scrub_explanation(tmpl.get("purpose") or "Spreadsheet valuation / ratio template"),
            inputs=list(tmpl.get("inputs") or [])[:12],
            outputs=list(tmpl.get("outputs") or [])[:12],
            decision_logic=[
                "Map model inputs to company evidence",
                "Trace formula dependencies without storing workbook cells",
                "Translate outputs into investment implications in AGI language",
            ],
            applications=["valuation", "scenario analysis", "sensitivity"],
            academy=classify_academy(str(tmpl.get("name") or "valuation")),
            source_book_id=book_id,
            source_chapter=tmpl.get("sheet"),
            confidence=0.72,
        )
        store.upsert_framework(fw)
        frameworks.append(fw)

    store.upsert_book(meta)
    rebuild_graph(store)
    store.record_spreadsheet(book_id)

    return {
        "ok": True,
        "kind": "spreadsheet",
        "book_id": book_id,
        "title": meta.title,
        "format": meta.source_format,
        "sheets": parsed.get("sheets") or [],
        "pages_processed": len(parsed.get("sheets") or []),
        "concepts_extracted": len(concepts),
        "frameworks_extracted": len(frameworks),
        "formulas_extracted": len(formulas),
        "knowledge_objects_created": len(concepts) + len(frameworks) + len(formulas),
        "academies": meta.academies,
        "variables": len(parsed.get("variables") or []),
        "named_ranges": len(parsed.get("named_ranges") or []),
        "raw_text_retained": False,
        "extraction_quality": _quality(len(formulas), len(concepts), len(frameworks)),
    }


def parse_spreadsheet(*, filename: str, content_bytes: bytes) -> dict[str, Any]:
    name = (filename or "").lower()
    if name.endswith(".csv"):
        return _parse_csv(content_bytes)
    if name.endswith((".xlsx", ".xlsm")):
        return _parse_xlsx(content_bytes)
    if name.endswith(".xls"):
        return _parse_xls(content_bytes)
    if name.endswith(".ods"):
        return _parse_ods(content_bytes)
    return {"ok": False, "reason": f"unsupported_sheet:{name}"}


def _parse_xlsx(raw: bytes) -> dict[str, Any]:
    try:
        import openpyxl
    except Exception:
        # fallback: pandas values only
        return _parse_via_pandas(raw, fmt="xlsx")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=False, read_only=True)
    except Exception as exc:
        return {"ok": False, "reason": f"xlsx_open_failed:{exc}"}

    sheets = list(wb.sheetnames)
    formulas: list[dict[str, Any]] = []
    variables: list[dict[str, Any]] = []
    templates: list[dict[str, Any]] = []
    named_ranges: list[str] = []
    topics: set[str] = set()
    academies: set[str] = set()

    try:
        for dn in getattr(wb, "defined_names", []) or []:
            named_ranges.append(str(dn))
    except Exception:
        pass

    for sheet_name in sheets[:20]:
        ws = wb[sheet_name]
        # Skip chart-only sheets (Damodaran workbooks often include Chartsheets).
        if not hasattr(ws, "iter_rows"):
            continue
        headers: list[str] = []
        formula_count = 0
        sample_inputs: list[str] = []
        sample_outputs: list[str] = []
        # read limited grid
        try:
            rows_iter = ws.iter_rows(min_row=1, max_row=80, max_col=24, values_only=False)
        except Exception:
            continue
        for r_i, row in enumerate(rows_iter):
            for cell in row:
                val = cell.value
                if val is None:
                    continue
                if r_i == 0 and isinstance(val, str) and val.strip():
                    headers.append(val.strip()[:60])
                if isinstance(val, str) and _FORMULA_CELL_RE.match(val.strip()):
                    formula_count += 1
                    expr = val.strip()[:160]
                    label = _infer_label(headers, cell.column_letter if hasattr(cell, "column_letter") else "", sheet_name)
                    hit = _match_ratio(f"{label} {expr} {sheet_name}")
                    if hit:
                        name, expression, meaning = hit
                        formulas.append(
                            {
                                "name": name,
                                "expression": expression if len(expr) < 8 else expr,
                                "meaning": meaning,
                                "sheet": sheet_name,
                                "confidence": 0.8,
                                "use_cases": ["spreadsheet model"],
                            }
                        )
                        topics.add(name.lower())
                        academies.add(classify_academy(name))
                        sample_outputs.append(name)
                    elif formula_count <= 12:
                        formulas.append(
                            {
                                "name": label or f"{sheet_name}_calc",
                                "expression": expr,
                                "meaning": "Spreadsheet calculated field",
                                "sheet": sheet_name,
                                "confidence": 0.65,
                            }
                        )
                        sample_outputs.append(label or sheet_name)
                elif isinstance(val, str) and val.strip() and r_i < 25:
                    token = val.strip()
                    if _looks_like_variable(token):
                        variables.append({"name": token[:60], "sheet": sheet_name, "meaning": f"Input/label on {sheet_name}"})
                        sample_inputs.append(token[:40])
                        hit = _match_ratio(token)
                        if hit:
                            name, expression, meaning = hit
                            formulas.append(
                                {
                                    "name": name,
                                    "expression": expression,
                                    "meaning": meaning,
                                    "sheet": sheet_name,
                                    "confidence": 0.78,
                                }
                            )
                            academies.add(classify_academy(name))

        if formula_count or headers:
            templates.append(
                {
                    "name": sheet_name,
                    "sheet": sheet_name,
                    "purpose": f"Spreadsheet template sheet '{sheet_name}'",
                    "inputs": sample_inputs[:10] or headers[:8],
                    "outputs": sample_outputs[:10],
                }
            )
            topics.add(sheet_name.lower()[:40])

    try:
        wb.close()
    except Exception:
        pass

    # de-dupe formulas by name
    uniq: dict[str, dict[str, Any]] = {}
    for f in formulas:
        uniq[str(f.get("name")).lower()] = f

    return {
        "ok": True,
        "format": "xlsx",
        "sheets": sheets,
        "formulas": list(uniq.values())[:40],
        "variables": _dedupe_vars(variables)[:40],
        "templates": templates[:20],
        "named_ranges": named_ranges[:40],
        "topics": sorted(topics)[:16],
        "academies": sorted(academies) or ["valuation"],
    }


def _parse_xls(raw: bytes) -> dict[str, Any]:
    return _parse_via_pandas(raw, fmt="xls")


def _parse_ods(raw: bytes) -> dict[str, Any]:
    return _parse_via_pandas(raw, fmt="ods")


def _parse_csv(raw: bytes) -> dict[str, Any]:
    try:
        text = raw.decode("utf-8")
    except UnicodeDecodeError:
        text = raw.decode("latin-1", errors="ignore")
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()][:80]
    headers = [h.strip() for h in re.split(r"[,;\t]", lines[0])] if lines else []
    variables = [{"name": h, "sheet": "csv", "meaning": "CSV column"} for h in headers if h]
    formulas = []
    tokens: list[str] = list(headers)
    for ln in lines[1:]:
        tokens.extend(x.strip() for x in re.split(r"[,;\t]", ln) if x.strip())
    for token in tokens:
        hit = _match_ratio(token)
        if hit:
            name, expression, meaning = hit
            if not any(f.get("name") == name for f in formulas):
                formulas.append(
                    {
                        "name": name,
                        "expression": expression,
                        "meaning": meaning,
                        "sheet": "csv",
                        "confidence": 0.7,
                    }
                )
        if _looks_like_variable(token) and not token.replace(".", "", 1).isdigit():
            variables.append({"name": token[:60], "sheet": "csv", "meaning": "CSV field"})
    return {
        "ok": True,
        "format": "csv",
        "sheets": ["csv"],
        "formulas": formulas,
        "variables": _dedupe_vars(variables)[:40],
        "templates": [
            {
                "name": "CSV Dataset",
                "sheet": "csv",
                "purpose": "Tabular dataset",
                "inputs": headers[:12],
                "outputs": [f["name"] for f in formulas][:12],
            }
        ],
        "named_ranges": [],
        "topics": [h.lower() for h in headers[:12]],
        "academies": sorted({classify_academy(f["name"]) for f in formulas}) or ["investment"],
    }


def _parse_via_pandas(raw: bytes, *, fmt: str) -> dict[str, Any]:
    try:
        import pandas as pd
    except Exception:
        return {"ok": False, "reason": "pandas_unavailable"}
    try:
        if fmt == "csv":
            df = pd.read_csv(io.BytesIO(raw))
            sheets = {"data": df}
        else:
            sheets = pd.read_excel(io.BytesIO(raw), sheet_name=None, engine=None)
    except Exception as exc:
        return {"ok": False, "reason": f"pandas_parse_failed:{exc}"}

    formulas: list[dict[str, Any]] = []
    variables: list[dict[str, Any]] = []
    templates: list[dict[str, Any]] = []
    for sheet_name, df in list((sheets or {}).items())[:15]:
        cols = [str(c) for c in list(df.columns)[:24]]
        for c in cols:
            variables.append({"name": c[:60], "sheet": str(sheet_name), "meaning": "Column variable"})
            hit = _match_ratio(c)
            if hit:
                name, expression, meaning = hit
                formulas.append({"name": name, "expression": expression, "meaning": meaning, "sheet": str(sheet_name), "confidence": 0.7})
        templates.append(
            {
                "name": str(sheet_name),
                "sheet": str(sheet_name),
                "purpose": f"{fmt.upper()} sheet template",
                "inputs": cols[:10],
                "outputs": [f["name"] for f in formulas if f.get("sheet") == str(sheet_name)][:10],
            }
        )
    return {
        "ok": True,
        "format": fmt,
        "sheets": list((sheets or {}).keys()),
        "formulas": formulas[:40],
        "variables": _dedupe_vars(variables)[:40],
        "templates": templates[:20],
        "named_ranges": [],
        "topics": [v["name"].lower() for v in variables[:12]],
        "academies": sorted({classify_academy(f["name"]) for f in formulas}) or ["valuation"],
    }


def _match_ratio(text: str) -> tuple[str, str, str] | None:
    blob = (text or "").lower().replace(" ", "")
    for key, trip in _RATIO_HINTS.items():
        if key.replace("/", "") in blob or key in (text or "").lower():
            return trip
    return None


def _looks_like_variable(token: str) -> bool:
    if len(token) < 3 or len(token) > 48:
        return False
    if token.startswith("="):
        return False
    return bool(re.search(r"[A-Za-z]", token)) and not token.isdigit()


def _infer_label(headers: list[str], col: str, sheet: str) -> str:
    if headers:
        return headers[0][:40]
    return f"{sheet}_{col}"[:40]


def _dedupe_vars(variables: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: set[str] = set()
    out = []
    for v in variables:
        k = str(v.get("name") or "").lower()
        if not k or k in seen:
            continue
        seen.add(k)
        out.append(v)
    return out


def _cid(prefix: str, *parts: str) -> str:
    h = hashlib.sha1("|".join(parts).encode("utf-8")).hexdigest()[:10]
    slug = re.sub(r"[^a-z0-9]+", "_", parts[0].lower()).strip("_")[:36]
    return f"{prefix}_{slug}_{h}"


def _sheet_id(name: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "_", (name or "sheet").lower()).strip("_")[:48]
    return f"sheet_{slug}"


def Path_stem(filename: str) -> str:
    from pathlib import Path

    return Path(filename).stem.replace("_", " ").strip() or "Spreadsheet"


def _quality(formulas: int, concepts: int, frameworks: int) -> str:
    score = formulas * 2 + concepts + frameworks
    if score >= 12:
        return "high"
    if score >= 5:
        return "medium"
    if score >= 1:
        return "low"
    return "empty"
