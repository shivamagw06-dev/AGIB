"""Controlled import of the repository's annual Capital IQ workbook.

The workbook is a source snapshot, not a live vendor call. It imports the full
2016–2026 supplied history and records the exact workbook and unit on every
write. This protects downstream ratio calculations from legacy Yahoo /
unknown-unit statement rows.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable

from financial_warehouse_completion.models import ENGINE_CODE, PROGRAMME_VERSION


SOURCE = "capital_iq_workbook"
DEFAULT_YEARS = tuple(range(2016, 2027))
WORKBOOK_PATH = Path(__file__).resolve().parents[2] / "2016-2026.xlsx"

FIELD_MAP = {
    "Revenue": "revenue", "Gross Profit": "gross_profit", "EBITDA": "ebitda",
    "EBIT": "ebit", "PBT": "pbt", "PAT": "pat", "EPS": "eps",
    "Cash & Equivalents": "cash", "Total Current Assets": "current_assets",
    "Total Assets": "assets", "Total Current Liabilities": "current_liabilities",
    "Total Debt": "debt", "Total Equity": "equity",
    "Working Capital": "working_capital", "Cash Flow from Operations": "cfo",
    "Capital Expenditure": "capex", "Free Cash Flow": "free_cash_flow",
    "Cash Flow from Investing": "cfi", "Cash Flow from Financing": "cff",
}


def _symbol(value: Any) -> str:
    text = str(value or "").strip().upper()
    return text.split(":", 1)[-1] if ":" in text else text


def _number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _sheet_rows(year: int, *, path: Path) -> Iterable[dict[str, Any]]:
    from openpyxl import load_workbook

    book = load_workbook(path, read_only=True, data_only=True)
    try:
        if str(year) not in book.sheetnames:
            return
        sheet = book[str(year)]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=3, max_row=3))]
        positions = {str(label).strip(): index for index, label in enumerate(headers) if label}
        for values in sheet.iter_rows(min_row=4, values_only=True):
            symbol = _symbol(values[positions.get("Ticker", 0)] if values else None)
            if not symbol:
                continue
            row: dict[str, Any] = {
                "symbol": symbol,
                "fiscal_year": f"FY{year}",
                "statement_type": "CONSOLIDATED",
                "statement_frequency": "ANNUAL",
                "source": SOURCE,
                "statement_version": f"capiq_workbook_{year}",
            }
            for label, field in FIELD_MAP.items():
                index = positions.get(label)
                if index is not None:
                    number = _number(values[index])
                    if number is not None:
                        row[field] = number
            # A row with no financial facts is a vendor no-coverage marker.
            if any(row.get(field) is not None for field in FIELD_MAP.values()):
                yield row
    finally:
        book.close()


def preview(*, years: Iterable[int] = DEFAULT_YEARS, path: Path = WORKBOOK_PATH) -> dict[str, Any]:
    if not path.is_file():
        return {"ok": False, "error": f"workbook_not_found:{path.name}"}
    summary: dict[str, int] = {}
    for year in years:
        summary[str(year)] = sum(1 for _ in _sheet_rows(int(year), path=path))
    return {
        "ok": True, "source": SOURCE, "workbook": path.name,
        "unit": "INR million", "years": summary,
        "engine": ENGINE_CODE, "version": PROGRAMME_VERSION,
    }


def audit_preview(*, years: Iterable[int] = DEFAULT_YEARS, path: Path = WORKBOOK_PATH,
                  limit: int | None = None) -> dict[str, Any]:
    """Run identity, period and completeness gates without writing financials."""
    from financial_warehouse_completion.capiq_normalization import audit_and_prepare

    selected = tuple(sorted({int(year) for year in years if 2016 <= int(year) <= 2026}))
    rows = [row for year in selected for row in _sheet_rows(year, path=path)]
    if limit is not None:
        rows = rows[:max(0, int(limit))]
    prepared = audit_and_prepare(rows, field_map=FIELD_MAP, source_file=path.name)
    audits = prepared["audits"]
    by_status: dict[str, int] = {}
    for row in audits:
        status = str(row.get("overall_status") or "UNKNOWN")
        by_status[status] = by_status.get(status, 0) + 1
    return {
        "ok": True, "source": SOURCE, "workbook": path.name, "years": list(selected),
        "seen": len(rows), "ready": len(prepared["accepted"]), "status_counts": by_status,
        "sample": audits[:25], "mapping_version": "CAPIQ_V1", "unit": "INR million",
    }


def import_completed_years(*, years: Iterable[int] = DEFAULT_YEARS, actor: str = "fwcp") -> dict[str, Any]:
    from institutional_warehouse.formulas import recalculate
    from financial_warehouse_completion.capiq_normalization import audit_and_prepare, persist

    selected = tuple(sorted({int(year) for year in years if 2016 <= int(year) <= 2026}))
    check = preview(years=selected)
    if not check.get("ok"):
        return check
    rows = [row for year in selected for row in _sheet_rows(year, path=WORKBOOK_PATH)]
    if not rows:
        return {"ok": False, "error": "no_financial_rows", **check}
    prepared = audit_and_prepare(rows, field_map=FIELD_MAP, source_file=WORKBOOK_PATH.name)
    written = persist(prepared, field_map=FIELD_MAP, actor=actor, source_file=WORKBOOK_PATH.name)
    rebuilt = recalculate(
        actor=actor,
        stages=("statement_derivations", "ratios", "annual_sector_ratios", "valuation", "factors", "quality"),
    )
    return {
        "ok": True, "source": SOURCE, "workbook": WORKBOOK_PATH.name,
        "years": list(selected), "rows": len(rows), "ready": len(prepared["accepted"]),
        "financials_annual": written["financials"], "identity": written["identity"],
        "audit": written["audit"], "metric_mapping": written["mapping"],
        "recalculated": rebuilt, "unit": "INR million", "engine": ENGINE_CODE,
        "version": PROGRAMME_VERSION,
    }
