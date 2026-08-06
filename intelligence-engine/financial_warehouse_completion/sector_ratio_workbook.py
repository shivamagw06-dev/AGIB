"""Controlled import of the checked-in Capital IQ sector-ratio workbook.

The file contains company-level, sector-specific historical ratios for FY2016
through FY2025.  It is deliberately separate from AGI's formula-derived
``annual_sector_ratios`` table: this module preserves vendor evidence first,
then derives quality-controlled annual medians for historical context.
"""

from __future__ import annotations

import re
import statistics
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Any, Iterable

from financial_warehouse_completion.models import ENGINE_CODE, PROGRAMME_VERSION


SOURCE = "capital_iq_sector_ratios"
SOURCE_VERSION = "sector_ratios_workbook_2026-08-06"
WORKBOOK_PATH = Path(__file__).resolve().parents[2] / "Sector ratios.xlsx"
_HEADER = re.compile(r"^(20\d{2})\s+(.+)$")

# Workbook tabs are specialist peer groups.  Map them to the sectors used by
# AGI's company master so a company's live ratio can be compared to its own
# historical peer set without mixing financially different businesses.
SECTOR_MAP = {
    "Banks": "Financials", "NBFC_Finance": "Financials", "Insurance": "Financials",
    "Fin_Markets": "Financials", "IT": "Information Technology", "Healthcare": "Health Care",
    "FMCG": "Consumer Staples", "Auto": "Consumer Discretionary",
    "Capital_Goods": "Industrials", "Chemicals": "Materials", "Construction": "Industrials",
    "Cement_Building": "Materials", "Consumer_Durables": "Consumer Discretionary",
    "Consumer_Services": "Consumer Discretionary", "Diversified": "Industrials",
    "Media_Entertainment": "Communication Services", "Metals_Mining": "Materials",
    "Oil_Gas": "Energy", "Power": "Utilities", "Realty": "Real Estate",
    "Services": "Industrials", "Telecom": "Communication Services", "Textiles": "Consumer Discretionary",
}

METRICS = {
    "ROE": ("roe", "PAT / average total equity"),
    "ROA": ("roa", "PAT / average total assets"),
    "P/E": ("pe", "Market cap at fiscal year-end / PAT"),
    "P/BV": ("pb", "Market cap at fiscal year-end / total equity"),
    "P/TBV": ("ptbv", "Market cap at fiscal year-end / tangible book value"),
    "P/Assets": ("p_assets", "Market cap at fiscal year-end / total assets"),
    "EV/EBITDA": ("ev_ebitda", "Enterprise value / EBITDA"),
    "EV/Sales": ("ev_sales", "Enterprise value / revenue"),
    "Net Debt/EBITDA": ("net_debt_ebitda", "Net debt / EBITDA"),
    "Debt/Equity": ("debt_equity", "Total debt / total equity"),
    "Net Debt/Equity": ("net_debt_equity", "Net debt / total equity"),
    "EBITDA Margin": ("ebitda_margin", "EBITDA / revenue"),
    "Gross Margin": ("gross_margin", "Gross profit / revenue"),
    "FCF Yield": ("fcf_yield", "Free cash flow / market capitalisation"),
    "R&D % Sales": ("rd_sales", "Research and development / revenue"),
}

_MULTIPLES = {"pe", "pb", "ptbv", "p_assets", "ev_ebitda", "ev_sales"}
_MEDIAN_CAP = {"pe": 300.0, "pb": 60.0, "ptbv": 60.0, "p_assets": 30.0,
               "ev_ebitda": 150.0, "ev_sales": 60.0}


def _number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        out = float(value)
    except (TypeError, ValueError):
        return None
    return out if out == out and abs(out) != float("inf") else None


def _symbol(value: Any) -> str:
    text = str(value or "").strip().upper()
    return text.split(":", 1)[-1] if ":" in text else text


def _fiscal_as_of(year: int) -> str:
    """Indian fiscal years in the supplied book close on 31 March."""
    return date(year, 3, 31).isoformat()


def _median_status(metric: str, value: float) -> tuple[str, str]:
    if metric in _MULTIPLES and value <= 0:
        return "EXCLUDED", "Non-positive multiple is not comparable for a sector median."
    cap = _MEDIAN_CAP.get(metric)
    if cap is not None and abs(value) > cap:
        return "EXCLUDED", f"Absolute value exceeds the {cap:g}x median-quality threshold."
    return "ELIGIBLE", "Reported Capital IQ value retained for historical peer analysis."


def _rows(*, path: Path = WORKBOOK_PATH) -> Iterable[dict[str, Any]]:
    from openpyxl import load_workbook

    book = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in book.worksheets:
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=4, max_row=4))]
            metric_columns: list[tuple[int, int, str]] = []
            for index, header in enumerate(headers):
                match = _HEADER.match(str(header or "").strip())
                if not match:
                    continue
                mapped = METRICS.get(match.group(2).strip())
                if mapped:
                    metric_columns.append((index, int(match.group(1)), mapped[0]))
            sector = SECTOR_MAP.get(sheet.title, sheet.title.replace("_", " "))
            for values in sheet.iter_rows(min_row=5, values_only=True):
                capiq_ticker = str(values[2] or "").strip() if len(values) > 2 else ""
                if not capiq_ticker.upper().startswith("NSEI:"):
                    continue
                symbol = _symbol(values[1] if len(values) > 1 else capiq_ticker)
                if not symbol:
                    continue
                company_name = str(values[0] or "").strip()
                for index, year, metric in metric_columns:
                    value = _number(values[index] if index < len(values) else None)
                    if value is None:
                        continue
                    eligibility, note = _median_status(metric, value)
                    yield {
                        "symbol": symbol, "fiscal_year": f"FY{year}", "metric": metric,
                        "source_version": SOURCE_VERSION, "as_of": _fiscal_as_of(year),
                        "company_name": company_name, "capiq_ticker": capiq_ticker,
                        "sector": sector, "source_sector": sheet.title, "value": value,
                        "definition": METRICS[next(name for name, item in METRICS.items() if item[0] == metric)][1],
                        "median_eligibility": eligibility, "quality_note": note,
                    }
    finally:
        book.close()


def preview(*, path: Path = WORKBOOK_PATH) -> dict[str, Any]:
    if not path.is_file():
        return {"ok": False, "error": f"workbook_not_found:{path.name}"}
    rows = list(_rows(path=path))
    return {
        "ok": True, "source": SOURCE, "source_version": SOURCE_VERSION,
        "workbook": path.name, "rows": len(rows),
        "companies": len({row["symbol"] for row in rows}),
        "years": sorted({row["fiscal_year"] for row in rows}),
        "metrics": sorted({row["metric"] for row in rows}),
        "eligible_for_medians": sum(row["median_eligibility"] == "ELIGIBLE" for row in rows),
        "engine": ENGINE_CODE, "version": PROGRAMME_VERSION,
    }


def import_history(*, actor: str = "fwcp", path: Path = WORKBOOK_PATH) -> dict[str, Any]:
    from institutional_warehouse import gateway

    check = preview(path=path)
    if not check.get("ok"):
        return check
    rows = list(_rows(path=path))
    raw = gateway.write(
        "sector_ratio_history", rows, source=SOURCE, actor=actor,
        reason="capital_iq_sector_ratios:authoritative_historical_baseline",
    )

    # Publish annual sector medians into the existing history table, which is
    # already consumed by valuation intelligence. Raw source rows stay intact.
    grouped: dict[tuple[str, str, str], list[float]] = defaultdict(list)
    for row in rows:
        if row["median_eligibility"] == "ELIGIBLE":
            grouped[(row["sector"], row["metric"], row["as_of"])].append(float(row["value"]))
    median_rows = [
        {"sector": sector, "metric": metric, "as_of": as_of,
         "median_value": round(statistics.median(values), 6), "company_count": len(values)}
        for (sector, metric, as_of), values in grouped.items() if len(values) >= 5
    ]
    medians = gateway.write(
        "historical_sector_medians", median_rows, source=SOURCE, actor=actor,
        reason="capital_iq_sector_ratios:annual_historical_medians",
    )
    return {"ok": bool(raw.get("ok") and medians.get("ok")), **check,
            "sector_ratio_history": raw, "historical_sector_medians": medians,
            "median_rows": len(median_rows)}
